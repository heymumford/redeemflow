"""Update partner Excel with Sprint R2 platform evaluation findings."""

from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

EXCEL_PATH = "RedeemFlow_Loyalty_Programs_Research.xlsx"

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BLUE = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")


def add_platform_evaluation_sheet(wb):
    ws = wb.create_sheet("Platform Evaluation")

    headers = [
        "Platform",
        "Tier",
        "Function",
        "API Access",
        "Auth Method",
        "Sandbox",
        "Pricing",
        "Coverage",
        "Startup-Friendly",
        "Decision",
        "Next Action",
        "Contact",
        "API Docs URL",
        "Key Risk",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = BLUE

    platforms = [
        [
            "Duffel",
            "T1: Search",
            "Revenue flight search + booking (300+ airlines via NDC/GDS/LCC)",
            "Public, self-service signup",
            "Bearer token",
            "Yes (Duffel Airways ZZ)",
            "$3/booking + 1% managed + $0.005/excess search",
            "300+ airlines (revenue only, no award flights)",
            "Yes",
            "SELECTED (Sprint R2)",
            "Sign up, build adapter, test sandbox",
            "duffel.com (self-service)",
            "duffel.com/docs/api",
            "Python SDK archived; use httpx directly",
        ],
        [
            "Currency Alliance",
            "T2: Commerce",
            "Point exchange, earn, burn, gift cards (~35 endpoints)",
            "Public docs, registration required for keys",
            "HMAC-SHA256",
            "Yes (sandbox.api.currencyalliance.com)",
            "2% of loyalty value transacted, no setup/minimums",
            "13+ airlines (Flying Blue, BA, Qatar, SQ, etc.), 1 hotel (Melia)",
            "Yes",
            "SELECTED for MVP (Sprint R2)",
            "Register, contact Chuck Ehredt (CEO) for platform partner terms",
            "currencyalliance.com/contact-us",
            "api.currencyalliance.com/api-docs/v3",
            "Limited US carrier coverage; Barcelona startup (vendor risk)",
        ],
        [
            "Points.com / Plusgrade",
            "T2: Commerce",
            "Buy/sell/transfer/exchange points (60+ programs)",
            "Enterprise-only, password-protected docs",
            "OAuth 2.0 MAC (HMAC-SHA256)",
            "No public sandbox",
            "Revenue share (principal model, ~14% margin on wholesale)",
            "60+ programs incl. all major US airlines + hotels",
            "No",
            "DEFERRED to scale phase",
            "Email partnerships@points.com (Sacha Diab, SVP Global BD)",
            "Sacha Diab (BD), Cara Sanna (Partner Strategy)",
            "points.readme.io (password-protected)",
            "Enterprise sales cycle (3-6 months); no self-service",
        ],
        [
            "MX Technologies",
            "T0: Read",
            "Credit card reward balances (Chase UR, Amex MR, etc.)",
            "Free developer account",
            "Basic auth (client_id + api_key)",
            "Yes (int-api.mx.com, test institutions)",
            "~$15K/yr average",
            "16,000+ financial institutions (credit card rewards only)",
            "Moderate",
            "EVALUATE (Sprint R3)",
            "Sign up at dashboard.mx.com, test rewards endpoints",
            "dashboard.mx.com/sign_up",
            "docs.mx.com",
            "Rewards endpoints are beta; does NOT cover airline/hotel loyalty",
        ],
        [
            "Yodlee (Envestnet)",
            "T0: Read",
            "Credit card reward balances (reward container)",
            "Developer registration required",
            "OAuth 2.0 (FastLink widget)",
            "Yes (5 pre-configured test users)",
            "$5-15K/mo platform + $0.10-$0.50/user/mo",
            "17,000+ financial institutions (credit card rewards only)",
            "No",
            "NOT SELECTED (pricing prohibitive)",
            "N/A",
            "developer.yodlee.com/user/register",
            "developer.yodlee.com",
            "Pricing prohibitive for startup; MX is cheaper alternative",
        ],
        [
            "AwardWallet (existing)",
            "T0: Read",
            "Standalone loyalty program balances (airlines, hotels, etc.)",
            "Partner API key",
            "Bearer token",
            "No (live API only)",
            "$4/user/month (managed)",
            "700+ loyalty programs (3 blocked: UA, AA, DL)",
            "Yes",
            "KEEP (no replacement available)",
            "Maintain existing adapter",
            "awardwallet.com/api",
            "awardwallet.com/api",
            "Blocked by 3 major US airlines; credential-scraping legal risk",
        ],
        [
            "Seats.aero (existing)",
            "T1: Search",
            "Award flight availability search",
            "Partner API key",
            "Partner-Authorization header",
            "No",
            "Subscription-based",
            "Major airline alliances (award availability)",
            "Yes",
            "KEEP (no alternative exists)",
            "Maintain existing adapter",
            "seats.aero",
            "seats.aero (partner docs)",
            "Air Canada lawsuit (CFAA + trademark, active)",
        ],
        [
            "Change API (existing)",
            "T3: Donate",
            "Point-to-charity donation processing",
            "Partner API key",
            "Bearer token",
            "Yes",
            "Per-transaction",
            "1.5M+ nonprofits",
            "Yes",
            "KEEP",
            "Maintain existing adapter",
            "getchange.io",
            "docs.getchange.io",
            "Money transmitter licensing risk for point-to-dollar flows",
        ],
        [
            "Stripe (existing)",
            "Billing",
            "Subscription management, checkout, webhooks",
            "Public, self-service",
            "API key (via SDK)",
            "Yes (test mode)",
            "2.9% + $0.30/transaction",
            "All subscription tiers",
            "Yes",
            "KEEP",
            "Maintain existing adapter",
            "stripe.com",
            "stripe.com/docs/api",
            "N/A",
        ],
    ]

    for row_idx, platform in enumerate(platforms, 2):
        for col_idx, value in enumerate(platform, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = WRAP

        decision = platform[9]
        decision_cell = ws.cell(row=row_idx, column=10)
        if "SELECTED" in decision:
            decision_cell.fill = GREEN
        elif "KEEP" in decision:
            decision_cell.fill = GREEN
        elif "EVALUATE" in decision:
            decision_cell.fill = YELLOW
        elif "DEFERRED" in decision:
            decision_cell.fill = YELLOW
        elif "NOT SELECTED" in decision:
            decision_cell.fill = RED

    col_widths = [22, 14, 50, 30, 22, 30, 40, 45, 14, 28, 45, 35, 35, 45]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    ws.auto_filter.ref = f"A1:N{len(platforms) + 1}"


def add_key_findings_sheet(wb):
    ws = wb.create_sheet("R2 Key Findings")

    headers = ["#", "Finding", "Impact", "Action Required"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = BLUE

    findings = [
        [
            1,
            "Yodlee and MX CANNOT read airline/hotel loyalty balances",
            "AwardWallet remains sole option for standalone loyalty programs. "
            "Yodlee/MX only read credit card rewards (Chase UR, Amex MR, etc.)",
            "Keep AwardWallet; add MX for credit card rewards as complement",
        ],
        [
            2,
            "No aggregator fills the United/Delta/American gap",
            "3 major US airlines actively block all third-party balance reads. "
            "Email parsing of monthly statements is the only workaround.",
            "Accept coverage gap; monitor CFPB Section 1033 for future relief",
        ],
        [
            3,
            "Amadeus Self-Service API shutting down July 2025",
            "Duffel becomes the clear choice for revenue flight search/booking. "
            "No viable alternative for NDC-first flight API.",
            "Prioritize Duffel integration",
        ],
        [
            4,
            "Currency Alliance has public API + sandbox (rare in loyalty space)",
            "Only T2 commerce platform accessible without enterprise sales. Points.com requires formal BD engagement.",
            "Build Currency Alliance adapter for MVP; pursue Points.com in parallel",
        ],
        [
            5,
            "Points.com docs are password-protected, API is enterprise-only",
            "Cannot evaluate or prototype without partnership agreement. Enterprise sales cycle expected (3-6 months).",
            "Email partnerships@points.com to begin conversation",
        ],
        [
            6,
            "CFPB Section 1033 includes 'rewards credits' as covered data",
            "If implemented, banks must share reward data via APIs. "
            "Benefits credit card rewards only, not airline/hotel programs.",
            "Monitor legal challenges; may reduce need for MX/Yodlee long-term",
        ],
        [
            7,
            "Duffel enables points-vs-cash comparison (new revenue stream)",
            "RedeemFlow can show award cost (Seats.aero) next to cash cost (Duffel). "
            "$3/booking revenue when users choose cash.",
            "Build Duffel adapter; add comparison view to search results",
        ],
        [
            8,
            "Currency Alliance coverage is Europe/Asia-heavy, weak on US carriers",
            "Flying Blue, BA, Qatar, Singapore confirmed but no United, Delta, American. "
            "Points.com has all major US carriers.",
            "Currency Alliance for MVP; Points.com partnership needed for US coverage",
        ],
        [
            9,
            "MX rewards endpoints are beta (API v20250224)",
            "Schema may change. Simpler data model than Yodlee's mature reward container.",
            "Build adapter with version pinning; expect breaking changes",
        ],
        [
            10,
            "Ascenda TransferConnect identified as potential alternative to Points.com",
            "50+ currencies, bank-focused (OCBC, Brex). "
            "May be more accessible than Points.com for cross-program exchange.",
            "Research Ascenda as fallback if Points.com partnership stalls",
        ],
    ]

    for row_idx, finding in enumerate(findings, 2):
        for col_idx, value in enumerate(finding, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = WRAP

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 55


def main():
    wb = load_workbook(EXCEL_PATH)

    for name in ["Platform Evaluation", "R2 Key Findings"]:
        if name in wb.sheetnames:
            del wb[name]

    add_platform_evaluation_sheet(wb)
    add_key_findings_sheet(wb)

    wb.save(EXCEL_PATH)
    print(f"Updated {EXCEL_PATH} with Platform Evaluation and R2 Key Findings sheets")
    print(f"Total sheets: {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        print(f"  - {name}")


if __name__ == "__main__":
    main()
