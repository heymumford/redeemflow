"""Update RedeemFlow partner Excel with enriched research data from 5 agents."""

from __future__ import annotations

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook("RedeemFlow_Loyalty_Programs_Research.xlsx")

# Style constants
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
RED_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
GREEN_FILL = PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
        cell.border = THIN_BORDER


def auto_width(ws, max_width=40):
    for col_cells in ws.columns:
        length = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 2, max_width)


def risk_fill(level: str):
    if "HIGH" in level.upper():
        return RED_FILL
    if "MEDIUM" in level.upper():
        return YELLOW_FILL
    return GREEN_FILL


# ── Sheet: Contacts ──────────────────────────────────────────────

if "Contacts" in wb.sheetnames:
    del wb["Contacts"]
ws = wb.create_sheet("Contacts", 0)

headers = [
    "Program",
    "Type",
    "Parent Company",
    "HQ",
    "Developer Portal URL",
    "Partnership Email",
    "Partnership Phone",
    "Affiliate Network",
    "Affiliate Apply URL",
    "Commission Rate",
    "Named Contact",
    "Notes",
]
ws.append(headers)
style_header(ws, len(headers))

contacts_data = [
    # Hotels
    [
        "Marriott Bonvoy",
        "Hotel",
        "Marriott International",
        "Bethesda, MD, USA",
        "https://devportalprod.marriott.com",
        None,
        "+1-301-380-3000",
        "Partnerize",
        "https://www.marriott.com/marriott/affiliateprogram.mi",
        "~2.7%",
        None,
        "Gated partner portal. Requires Marriott sponsorship.",
    ],
    [
        "Hilton Honors",
        "Hotel",
        "Hilton Worldwide Holdings",
        "McLean, VA, USA",
        "https://developer.hilton.io",
        "API_Request@hilton.com",
        "+1-703-883-1000",
        "Rakuten (Americas), Tradedoubler (EU), Awin (UK)",
        None,
        "~4%",
        None,
        "Most comprehensive hotel API docs. OAuth2 auth.",
    ],
    [
        "IHG One Rewards",
        "Hotel",
        "InterContinental Hotels Group",
        "Windsor, UK",
        None,
        None,
        "+44 (0)1753 972 000",
        "CJ Affiliate",
        "https://partnerconnect.ihg.com/",
        "3%",
        None,
        "PartnerConnect portal. 14-day cookie.",
    ],
    [
        "World of Hyatt",
        "Hotel",
        "Hyatt Hotels Corporation",
        "Chicago, IL, USA",
        None,
        None,
        "800-544-9288",
        "UNVERIFIED",
        None,
        None,
        None,
        "No developer portal. No public API.",
    ],
    [
        "Wyndham Rewards",
        "Hotel",
        "Wyndham Hotels & Resorts",
        "Parsippany, NJ, USA",
        None,
        None,
        None,
        "Points.com",
        "https://www.wyndhamhotels.com/hotel-deals/partner-program",
        "~2.5%",
        None,
        "Points.com confirmed for buy/sell.",
    ],
    [
        "Choice Privileges",
        "Hotel",
        "Choice Hotels International",
        "Rockville, MD, USA",
        None,
        None,
        "(301) 592-5000",
        "Direct",
        "https://www.choicehotels.com/affiliate",
        "2-6%",
        None,
        "MuleSoft internal APIs. No public portal.",
    ],
    [
        "Accor Live Limitless",
        "Hotel",
        "Accor SA",
        "Issy-les-Moulineaux, France",
        None,
        None,
        None,
        "CJ / Direct",
        "https://all.accor.com/a/en/affiliate-program.html",
        "Up to 8%",
        None,
        "Highest affiliate commission. No US co-branded card.",
    ],
    [
        "Radisson Rewards",
        "Hotel",
        "Radisson Hotel Group (Jin Jiang)",
        "Brussels, Belgium",
        None,
        None,
        None,
        "Awin",
        "https://ui.awin.com/merchant-profile-terms/5907",
        "6%",
        None,
        "Americas properties migrated to Choice. EU/global only.",
    ],
    # Airlines
    [
        "United MileagePlus",
        "Airline",
        "United Airlines Holdings",
        "Chicago, IL, USA",
        None,
        "customercare@united.com",
        "(800) 421-4655",
        None,
        None,
        None,
        None,
        "No developer portal. BLOCKED by AwardWallet.",
    ],
    [
        "American AAdvantage",
        "Airline",
        "American Airlines Group",
        "Fort Worth, TX, USA",
        "https://aa-dev-passenger.apigee.io",
        None,
        None,
        None,
        None,
        None,
        None,
        "Apigee portal gated. BLOCKED by AwardWallet.",
    ],
    [
        "Delta SkyMiles",
        "Airline",
        "Delta Air Lines",
        "Atlanta, GA, USA",
        None,
        None,
        None,
        None,
        None,
        None,
        "Dwight James (SVP Customer Engagement & Loyalty)",
        "No API. BLOCKED by AwardWallet. C&D history.",
    ],
    [
        "Southwest Rapid Rewards",
        "Airline",
        "Southwest Airlines Co.",
        "Dallas, TX, USA",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "SWABIZ for B2B. Points.com 9+ year partnership.",
    ],
    [
        "JetBlue TrueBlue",
        "Airline",
        "JetBlue Airways",
        "Long Island City, NY, USA",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "Most inbound bank partners (6). No developer portal.",
    ],
    [
        "Alaska Atmos Rewards",
        "Airline",
        "Alaska Air Group",
        "Seattle, WA, USA",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "Rebranded from Mileage Plan Oct 2025.",
    ],
    [
        "Air Canada Aeroplan",
        "Airline",
        "Air Canada",
        "Montreal, QC, Canada",
        None,
        None,
        "1-800-361-5373",
        None,
        None,
        None,
        None,
        "Broadest US bank transfer network. Sued Seats.aero.",
    ],
    [
        "British Airways Avios",
        "Airline",
        "IAG",
        "London, UK",
        "https://developer.iagloyalty.com",
        "customerrelations@avios.com",
        None,
        None,
        None,
        None,
        None,
        "BEST airline loyalty API. Balance/earn/redeem endpoints.",
    ],
    [
        "Singapore KrisFlyer",
        "Airline",
        "Singapore Airlines",
        "Singapore",
        "https://developer.singaporeair.com",
        "public_affairs@singaporeair.com.sg",
        None,
        None,
        None,
        None,
        None,
        "KrisConnect: OAuth, sandbox, earn/redeem APIs.",
    ],
    [
        "ANA Mileage Club",
        "Airline",
        "ANA Holdings",
        "Tokyo, Japan",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "Amex-exclusive for US transfers. No outbound transfers.",
    ],
    [
        "Emirates Skywards",
        "Airline",
        "Emirates Group",
        "Dubai, UAE",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "IBS iLoyal platform. Chase dropped Oct 2024.",
    ],
    [
        "Flying Blue",
        "Airline",
        "Air France-KLM Group",
        "Paris, France",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "Points.com confirmed. Broadest inbound network.",
    ],
    [
        "Turkish Miles&Smiles",
        "Airline",
        "Turkish Airlines",
        "Istanbul, Turkey",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "Devalued Dec 2025. Capital One/Citi/Bilt only.",
    ],
    [
        "Cathay Asia Miles",
        "Airline",
        "Cathay Pacific (Swire)",
        "Hong Kong",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "800+ partners. 48hr transfer hold.",
    ],
    # Bank Programs
    [
        "Chase Ultimate Rewards",
        "Bank",
        "JPMorgan Chase",
        "New York, NY, USA",
        "https://developer.chase.com",
        None,
        None,
        None,
        None,
        None,
        None,
        "Account data API only. No rewards balance API.",
    ],
    [
        "Amex Membership Rewards",
        "Bank",
        "American Express",
        "New York, NY, USA",
        "https://developer.americanexpress.com",
        None,
        None,
        None,
        None,
        None,
        None,
        "Smart Offer API. No rewards balance API.",
    ],
    [
        "Citi ThankYou Points",
        "Bank",
        "Citigroup",
        "New York, NY, USA",
        "https://developer.citi.com",
        None,
        None,
        None,
        None,
        None,
        None,
        "Pay with Points redemption API. No balance read.",
    ],
    [
        "Capital One Miles",
        "Bank",
        "Capital One Financial",
        "McLean, VA, USA",
        "https://developer.capitalone.com",
        None,
        None,
        None,
        None,
        None,
        None,
        "ONLY bank with public Rewards Balance API (DevExchange).",
    ],
    [
        "Bilt Rewards",
        "Bank",
        "Bilt Technologies",
        "New York, NY, USA",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "$10.75B valuation. Cardless/Column Bank issuer. No API.",
    ],
    [
        "Wells Fargo Rewards",
        "Bank",
        "Wells Fargo",
        "San Francisco, CA, USA",
        "https://developer.wellsfargo.com",
        None,
        None,
        None,
        None,
        None,
        None,
        "Newest Big 6 (April 2024). Commercial APIs only.",
    ],
    # Exchange Platforms
    [
        "Points.com (Plusgrade)",
        "Exchange",
        "Plusgrade",
        "Montreal, QC, Canada",
        "https://lcp.points.com/v1",
        None,
        "+1.514.437.2185",
        None,
        None,
        "Revenue share (negotiated)",
        "Cara Sanna (SVP Partner Strategy)",
        "60+ programs. 92B+ points/year. B2B enterprise only. No public sandbox.",
    ],
    [
        "Currency Alliance",
        "Exchange",
        "Currency Alliance",
        "London, UK",
        "https://api.currencyalliance.com/api-docs/v3",
        None,
        None,
        None,
        None,
        "SaaS + free integration for brands",
        None,
        "~30 API endpoints. IATA experience. Startup-friendly.",
    ],
    [
        "AwardWallet",
        "Aggregator",
        "AwardWallet",
        None,
        "https://awardwallet.com/api/main",
        None,
        None,
        None,
        None,
        "Custom/quote-based",
        None,
        "3 APIs: Email Parse, Web Parse, Account Access OAuth. 600+ programs.",
    ],
    [
        "Seats.aero",
        "Search",
        "Localhost",
        None,
        "https://developers.seats.aero",
        None,
        None,
        None,
        None,
        "$10/month Pro",
        None,
        "OAuth model. SUED by Air Canada (CFAA + trademark).",
    ],
]

for row in contacts_data:
    ws.append(row)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
    for cell in row:
        cell.alignment = WRAP
        cell.border = THIN_BORDER

auto_width(ws)

# ── Sheet: Legal Risk ────────────────────────────────────────────

if "Legal Risk" in wb.sheetnames:
    del wb["Legal Risk"]
ws = wb.create_sheet("Legal Risk", 1)

headers = [
    "Program",
    "Risk Level",
    "TOS Prohibits Scraping",
    "C&D History",
    "Lawsuit History",
    "AwardWallet Status",
    "Notes",
]
ws.append(headers)
style_header(ws, len(headers))

legal_data = [
    [
        "United MileagePlus",
        "HIGH",
        "Yes",
        "None confirmed",
        None,
        "BLOCKED (~2023)",
        "Written authorization required for commercial access.",
    ],
    [
        "American AAdvantage",
        "HIGH",
        "Yes — explicit anti-scraping clause",
        "C&D to AwardWallet, C&D to The Points Guy (Jan 2022)",
        "AA v. TPG (2022-2023, settled)",
        "BLOCKED",
        "60K miles/year transfer cap via Points.com.",
    ],
    [
        "Delta SkyMiles",
        "HIGH",
        "Yes — prohibits sale/barter/assignment",
        "C&D to AwardWallet (Sept 2012), MileWise, TripIt",
        None,
        "BLOCKED (since 2012)",
        "First program to C&D aggregators.",
    ],
    [
        "Air Canada Aeroplan",
        "HIGH",
        None,
        None,
        "Air Canada v. Seats.aero (Oct 2023, CFAA + trademark, active)",
        "Tracks",
        "Sued Seats.aero for $2M+ per service. Most aggressive enforcement.",
    ],
    ["Southwest Rapid Rewards", "LOW", "Standard terms", None, None, "Tracks", "9+ year Points.com partnership."],
    ["JetBlue TrueBlue", "LOW", "No specific clause found", None, None, "Tracks", None],
    ["Alaska Atmos Rewards", "LOW", "No specific clause found", None, None, "Tracks", None],
    [
        "British Airways Avios",
        "LOW",
        "API gated by partner registration",
        None,
        None,
        "Tracks",
        "IAG Developer Portal is the proper channel.",
    ],
    ["Singapore KrisFlyer", "LOW", "OAuth by design", None, None, "Tracks", "KrisConnect requires NDA for production."],
    ["Emirates Skywards", "LOW", "No specific clause found", None, None, "Tracks", None],
    [
        "Marriott Bonvoy",
        "MEDIUM",
        "Yes — prohibits robots/spiders/scraping",
        None,
        None,
        "Tracks",
        "Third-party bookings don't earn points. Commercial use requires consent.",
    ],
    ["Hilton Honors", "MEDIUM", "Standard terms", None, None, "Tracks", "API access requires partner approval."],
    [
        "ANA Mileage Club",
        "MEDIUM",
        "Prohibits outbound transfers entirely",
        None,
        None,
        "Tracks",
        "Cannot transfer miles to other programs.",
    ],
]

for row in legal_data:
    ws.append(row)
    # Color the risk level cell
    risk_cell = ws.cell(row=ws.max_row, column=2)
    risk_cell.fill = risk_fill(row[1])

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
    for cell in row:
        cell.alignment = WRAP
        cell.border = THIN_BORDER

auto_width(ws)

# ── Sheet: Transfer Matrix ───────────────────────────────────────

if "Transfer Matrix" in wb.sheetnames:
    del wb["Transfer Matrix"]
ws = wb.create_sheet("Transfer Matrix", 2)

headers = ["Program", "Amex MR", "Chase UR", "Capital One", "Citi TY", "Bilt", "Wells Fargo", "Total Inbound"]
ws.append(headers)
style_header(ws, len(headers))

transfer_data = [
    # Airlines
    ["United MileagePlus", "--", "1:1", "--", "--", "1:1", "--", 2],
    ["American AAdvantage", "--", "--", "--", "1:1", "--", "--", 1],
    ["Delta SkyMiles", "1:1", "--", "--", "--", "--", "--", 1],
    ["Southwest Rapid Rewards", "--", "1:1", "--", "--", "1:1", "--", 2],
    ["JetBlue TrueBlue", "1:0.8", "1:1", "5:3", "1:1", "--", "1:1", 5],
    ["Alaska Atmos Rewards", "--", "--", "--", "--", "1:1", "--", 1],
    ["Air Canada Aeroplan", "1:1", "1:1", "1:1", "1:1", "1:1", "--", 5],
    ["British Airways Avios", "1:1", "1:1", "1:1", "--", "1:1", "--", 4],
    ["Singapore KrisFlyer", "1:1", "1:1", "1:1", "1:1", "--", "--", 4],
    ["ANA Mileage Club", "1:1", "--", "--", "--", "--", "--", 1],
    ["Emirates Skywards", "1:0.8", "DROPPED", "devalued", "1:0.8", "1:1", "--", 3],
    ["Qantas Frequent Flyer", "1:1", "--", "1:1", "1:1", "--", "--", 3],
    ["Virgin Atlantic Flying Club", "1:1", "1:1", "1:1", "1:1", "1:1", "--", 5],
    ["Air France/KLM Flying Blue", "1:1", "1:1", "1:1", "1:1", "1:1", "1:1", 6],
    ["Turkish Miles&Smiles", "--", "--", "1:1", "1:1", "1:1", "--", 3],
    ["Cathay Pacific Asia Miles", "1:1", "--", "1:1", "1:1", "1:1", "--", 4],
    # Hotels
    ["Marriott Bonvoy", "1:1", "1:1", "--", "--", "1:1", "--", 3],
    ["Hilton Honors", "1:2", "--", "--", "--", "1:1", "--", 2],
    ["IHG One Rewards", "--", "1:1", "--", "--", "1:1", "--", 2],
    ["World of Hyatt", "--", "1:1", "--", "--", "1:1", "--", 2],
    ["Wyndham Rewards", "--", "1:1", "1:1", "1:1", "--", "--", 3],
    ["Choice Privileges", "--", "--", "1:1", "1:2", "--", "1:2", 2],
    ["Accor Live Limitless", "--", "--", "2:1", "2:1", "3:2", "--", 3],
]

for row in transfer_data:
    ws.append(row)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
    for cell in row:
        cell.alignment = WRAP
        cell.border = THIN_BORDER

auto_width(ws)

# ── Sheet: Integration Architecture ─────────────────────────────

if "Integration Architecture" in wb.sheetnames:
    del wb["Integration Architecture"]
ws = wb.create_sheet("Integration Architecture", 3)

headers = [
    "Tier",
    "Platform",
    "What It Does",
    "Status",
    "Programs Covered",
    "Pricing",
    "Contact",
    "Legal Risk",
    "Priority",
]
ws.append(headers)
style_header(ws, len(headers))

arch_data = [
    [
        "T0: Balance Read",
        "AwardWallet",
        "Screen-scrape 600+ loyalty balances",
        "INTEGRATED",
        "~600 (minus AA, Delta, United)",
        "Custom/quote",
        "awardwallet.com/contact.php",
        "MEDIUM (scraping model)",
        "P0",
    ],
    [
        "T0: Balance Read",
        "Capital One DevExchange",
        "Official rewards balance API",
        "NOT STARTED",
        "Capital One only",
        "Free (registration required)",
        "developer.capitalone.com",
        "LOW (sanctioned)",
        "P1",
    ],
    [
        "T0: Balance Read",
        "Yodlee Reward Container",
        "OAuth financial data + rewards",
        "NOT EVALUATED",
        "UNKNOWN — needs sandbox test",
        "Custom/quote",
        "developer.yodlee.com",
        "LOW (OAuth)",
        "P2",
    ],
    [
        "T1: Award Search",
        "Seats.aero",
        "Award availability across 24+ programs",
        "INTEGRATED",
        "24+ airlines",
        "$10/mo Pro",
        "developers.seats.aero",
        "MEDIUM (sued by Air Canada)",
        "P0",
    ],
    [
        "T1: Flight Booking",
        "Duffel",
        "NDC flight search/booking for 300+ airlines",
        "NOT STARTED",
        "300+ airlines",
        "Free tier + per-booking",
        "duffel.com",
        "LOW (sanctioned NDC)",
        "P2",
    ],
    [
        "T2: Commerce",
        "Points.com (Plusgrade)",
        "Buy/sell/transfer/exchange points",
        "NOT STARTED",
        "60+ programs",
        "Enterprise B2B (revenue share)",
        "plusgrade.com/contact, Cara Sanna (SVP)",
        "LOW (programs authorize)",
        "P1",
    ],
    [
        "T2: Commerce",
        "Currency Alliance",
        "Brand-to-brand point exchange",
        "NOT STARTED",
        "Variable",
        "SaaS + free brand integration",
        "api.currencyalliance.com",
        "LOW (authorized exchange)",
        "P1",
    ],
    [
        "T2: Commerce",
        "IAG Loyalty API",
        "Avios earn/redeem/balance (BA/IB/EI)",
        "NOT STARTED",
        "5 IAG airlines",
        "Partner registration",
        "developer.iagloyalty.com",
        "LOW (official API)",
        "P2",
    ],
    [
        "T2: Commerce",
        "SIA KrisConnect",
        "KrisFlyer earn/redeem/auth (OAuth)",
        "NOT STARTED",
        "Singapore Airlines",
        "Partner registration + NDA",
        "developer.singaporeair.com",
        "LOW (official API)",
        "P2",
    ],
    [
        "T3: Donate",
        "Change API",
        "Point-to-charity donation pipeline",
        "INTEGRATED",
        "50-state CCV compliance",
        "Per-transaction",
        "getchange.io",
        "LOW",
        "P0",
    ],
    [
        "T3: Donate",
        "Pledge.to",
        "2M+ nonprofit donation API",
        "NOT STARTED",
        "2M+ nonprofits",
        "Free API",
        "pledge.to",
        "LOW",
        "P3",
    ],
]

for row in arch_data:
    ws.append(row)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
    for cell in row:
        cell.alignment = WRAP
        cell.border = THIN_BORDER
    # Color legal risk
    risk_cell = ws.cell(row=row[0].row, column=8)
    risk_cell.fill = risk_fill(str(risk_cell.value or ""))

auto_width(ws)

# ── Sheet: Legal Actions Timeline ────────────────────────────────

if "Legal Actions" in wb.sheetnames:
    del wb["Legal Actions"]
ws = wb.create_sheet("Legal Actions", 4)

headers = ["Year", "Action", "Plaintiff", "Defendant", "Claims", "Outcome", "Relevance to RedeemFlow"]
ws.append(headers)
style_header(ws, len(headers))

legal_actions = [
    [
        2012,
        "C&D letter",
        "Delta SkyMiles",
        "AwardWallet",
        "Unauthorized automated access",
        "AwardWallet dropped SkyMiles support",
        "HIGH — same scraping model we use via AwardWallet",
    ],
    [
        2012,
        "C&D letter",
        "Delta SkyMiles",
        "MileWise, TripIt",
        "Unauthorized automated access",
        "Compliance",
        "MEDIUM — pattern of enforcement",
    ],
    [
        2021,
        "C&D letter",
        "American AAdvantage",
        "AwardWallet",
        "Unauthorized automated access",
        "AwardWallet stopped tracking AAdvantage",
        "HIGH — same model",
    ],
    [
        2022,
        "Federal lawsuit",
        "American Airlines",
        "The Points Guy",
        "Commercial use of trademarks, competitive harm",
        "Settled (terms undisclosed)",
        "MEDIUM — trademark risk for program names",
    ],
    [
        2023,
        "Federal lawsuit",
        "Air Canada (Aeroplan)",
        "Seats.aero (Localhost)",
        "CFAA (unauthorized access) + trademark infringement",
        "ACTIVE — seeking $2M+ per service type",
        "HIGH — we use Seats.aero. Association risk.",
    ],
]

for row in legal_actions:
    ws.append(row)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
    for cell in row:
        cell.alignment = WRAP
        cell.border = THIN_BORDER

auto_width(ws)

# Save
wb.save("RedeemFlow_Loyalty_Programs_Research.xlsx")
print("Excel updated: 5 new sheets (Contacts, Legal Risk, Transfer Matrix, Integration Architecture, Legal Actions)")
